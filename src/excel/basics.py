from openpyxl import Workbook
from openpyxl.formula.translate import Translator

book = Workbook()
new_sheet = book.active

for i in range(1, 10001):
    new_sheet['A' + str(i)] = i
    new_sheet['B' + str(i)] = i * i
    if i == 1:
        new_sheet['C' + str(i)] = "=A1*B1"
    else:
        new_sheet['C' + str(i)] = Translator("=A1*B1", origin="C1").translate_formula('C' + str(i))

print(f'Max rows: {new_sheet.max_row}, max columns: {new_sheet.max_column}')
book.save("file.xlsx")


